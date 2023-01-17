from pandas import date_range
import tweepy
import twitter_credentials as t
import datetime
from openpyxl import Workbook
import openpyxl
wb = openpyxl.load_workbook("twitter_results.xlsx")
ws = wb.active
dest_filename = "twitter_results.xlsx"

#Authorization to scrape Tweets
client = tweepy.Client(bearer_token=t.bearer_token,wait_on_rate_limit=True)
#Assigning values for the parameters
query = 'sinovac lang:tl -is:retweet'
    #[vacbrand] is subject to change based on the vaccine brand to be searched for
start_time= '2021-03-01T12:00:01Z'
end_time= '2021-12-31T11:59:59Z'
    #Values of start_time and end_time are subject to changes based on desired timeframe
count = 1
for tweet in tweepy.Paginator(client.search_all_tweets,query=query, start_time=start_time, end_time=end_time, max_results=500).flatten(limit=100000):
    ws.cell(row=count,column=1,value=str(count))
    ws.cell(row=count,column=2,value=str(tweet))
    count+=1
    wb.save(filename = dest_filename)

wb.save(filename = dest_filename)
